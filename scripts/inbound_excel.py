#!/usr/bin/env python
import argparse
import json
import re
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency: openpyxl. Install it with `python -m pip install openpyxl`."
    ) from exc


MAIN_SHEET_NAMES = ["项目需求填写界面"]

MAIN_HEADER_ALIASES = {
    "demandFactoryCode": ["需求工厂代码", "工厂代码"],
    "wbsCode": ["WBS编码", "WBS项目编码"],
    "projectDefinition": ["项目定义"],
    "materialCode": ["物料编码"],
    "materialName": ["物料名称"],
    "purchaseQuantity": ["需求数量", "采购数量", "数量"],
    "unit": ["基本计量单位", "采购单位", "单位"],
    "mrpController": ["MRP控制者", "MRP编码"],
}


def normalize_header(value):
    return re.sub(r"\s+", "", str(value or "")).strip()


def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, date):
        return value.strftime("%Y%m%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_decimal(value, field_name):
    text = normalize_cell(value)
    if not text:
        raise ValueError(f"{field_name} is empty.")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{field_name} must be numeric; got {text!r}.") from exc


def decimal_text(value):
    normalized = value.normalize()
    if normalized == normalized.to_integral():
        return str(normalized.quantize(Decimal(1)))
    return format(normalized, "f")


def find_sheet(workbook, names, fallback_index=None):
    for name in names:
        if name in workbook.sheetnames:
            return workbook[name]
    if fallback_index is not None and fallback_index < len(workbook.worksheets):
        return workbook.worksheets[fallback_index]
    raise ValueError(f"Workbook is missing sheet: {names[0]}")


def find_column(headers, aliases):
    normalized_aliases = {normalize_header(alias) for alias in aliases}
    for index, header in enumerate(headers, start=1):
        if normalize_header(header) in normalized_aliases:
            return index
    return None


def resolve_columns(ws, aliases_by_key):
    headers = [ws.cell(row=1, column=column).value for column in range(1, ws.max_column + 1)]
    columns = {}
    for key, aliases in aliases_by_key.items():
        column = find_column(headers, aliases)
        if column:
            columns[key] = column
    return columns, [normalize_cell(item) for item in headers]


def unique_non_empty(values, field_name):
    unique = sorted({value for value in values if value})
    if len(unique) != 1:
        raise ValueError(f"{field_name} must contain exactly one unique non-empty value; got {unique!r}.")
    return unique[0]


def data_row_indices(ws, start_row, columns):
    rows = []
    for row in range(start_row, ws.max_row + 1):
        values = [normalize_cell(ws.cell(row=row, column=column).value) for column in columns]
        if any(values):
            rows.append(row)
    return rows


def run(input_path):
    input_file = Path(input_path).expanduser().resolve()
    if not input_file.exists():
        raise FileNotFoundError(f"Excel file does not exist: {input_file}")

    wb = load_workbook(input_file, data_only=True)
    main_ws = find_sheet(wb, MAIN_SHEET_NAMES, fallback_index=0)
    columns, headers = resolve_columns(main_ws, MAIN_HEADER_ALIASES)
    required = ["demandFactoryCode", "wbsCode", "projectDefinition", "materialCode", "purchaseQuantity"]
    missing = [key for key in required if key not in columns]
    if missing:
        raise ValueError(f"Main sheet is missing required columns: {missing!r}.")

    rows = data_row_indices(main_ws, 3, [columns[key] for key in required])
    if not rows:
        raise ValueError("Main sheet contains no data rows from row 3 onward.")

    demand_factory_code = unique_non_empty(
        [normalize_cell(main_ws.cell(row=row, column=columns["demandFactoryCode"]).value) for row in rows],
        "需求工厂代码",
    )
    wbs_code = unique_non_empty(
        [normalize_cell(main_ws.cell(row=row, column=columns["wbsCode"]).value) for row in rows],
        "WBS编码",
    )
    project_definition = unique_non_empty(
        [normalize_cell(main_ws.cell(row=row, column=columns["projectDefinition"]).value) for row in rows],
        "项目定义",
    )

    material_rows = []
    quantity_totals = {}
    for row in rows:
        material_code = normalize_cell(main_ws.cell(row=row, column=columns["materialCode"]).value)
        if not material_code:
            raise ValueError(f"Row {row} is missing 物料编码.")
        quantity = parse_decimal(
            main_ws.cell(row=row, column=columns["purchaseQuantity"]).value,
            f"Row {row} 需求数量",
        )
        quantity_totals[material_code] = quantity_totals.get(material_code, Decimal(0)) + quantity
        material_rows.append({
            "row": row,
            "materialCode": material_code,
            "materialName": normalize_cell(main_ws.cell(row=row, column=columns.get("materialName", 0)).value) if columns.get("materialName") else "",
            "purchaseQuantity": decimal_text(quantity),
            "unit": normalize_cell(main_ws.cell(row=row, column=columns.get("unit", 0)).value) if columns.get("unit") else "",
        })

    return {
        "ok": True,
        "inputPath": str(input_file),
        "sheetName": main_ws.title,
        "headerRow": 1,
        "instructionRow": 2,
        "dataStartRow": 3,
        "dataRows": rows,
        "headers": headers,
        "columns": columns,
        "projectDefinition": project_definition,
        "wbsCode": wbs_code,
        "demandFactoryCode": demand_factory_code,
        "mrpController": normalize_cell(main_ws.cell(row=rows[0], column=columns.get("mrpController", 0)).value) if columns.get("mrpController") else "",
        "materialRows": material_rows,
        "quantityByMaterialCode": {
            material_code: decimal_text(quantity)
            for material_code, quantity in sorted(quantity_totals.items())
        },
    }


def main():
    parser = argparse.ArgumentParser(description="Extract workflow 414 inbound input data from a purchase workbook.")
    parser.add_argument("--input", required=True)
    args = parser.parse_args()
    try:
        result = run(args.input)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as error:
        print(json.dumps({"ok": False, "error": str(error)}, ensure_ascii=False, indent=2), file=sys.stderr)
        raise SystemExit(1)


if __name__ == "__main__":
    main()
