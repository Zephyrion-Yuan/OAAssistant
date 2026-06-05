#!/usr/bin/env python
import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency: openpyxl. Install it with `python -m pip install openpyxl`."
    ) from exc


MAIN_SHEET_NAMES = ["项目需求填写界面"]
MRP_SHEET_NAMES = ["MRP控制者"]
COST_CENTER_SHEET_NAMES = ["成本中心"]

MAIN_HEADER_ALIASES = {
    "demandFactoryCode": ["需求工厂代码", "工厂代码"],
    "wbsCode": ["WBS编码", "WBS项目编码"],
    "projectDefinition": ["项目定义"],
    "mrpController": ["MRP控制者", "MRP编码"],
    "materialCode": ["物料编码"],
    "materialName": ["物料名称"],
    "demandQuantity": ["需求数量"],
    "unit": ["基本计量单位", "采购单位", "单位"],
}

COST_HEADER_ALIASES = {
    "costCenterCode": ["成本中心"],
    "company": ["公司", "公司代码"],
    "shortText": ["短文本", "成本中心名称"],
}


def normalize_header(value):
    return re.sub(r"\s+", "", str(value or "")).strip()


def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, date):
        return value.strftime("%Y%m%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def find_sheet(workbook, names, fallback_index=None):
    for name in names:
        if name in workbook.sheetnames:
            return workbook[name]
    if fallback_index is not None and fallback_index < len(workbook.worksheets):
        return workbook.worksheets[fallback_index]
    raise ValueError(f"Workbook is missing sheet: {names[0]}")


def find_column(headers, aliases):
    normalized_aliases = {normalize_header(alias) for alias in aliases}
    for index, header in enumerate(headers, start=1):
        if normalize_header(header) in normalized_aliases:
            return index
    return None


def resolve_columns(ws, aliases_by_key):
    headers = [ws.cell(row=1, column=column).value for column in range(1, ws.max_column + 1)]
    columns = {}
    for key, aliases in aliases_by_key.items():
        column = find_column(headers, aliases)
        if column:
            columns[key] = column
    return columns, [normalize_cell(item) for item in headers]


def unique_non_empty(values, field_name):
    unique = sorted({value for value in values if value})
    if len(unique) != 1:
        raise ValueError(f"{field_name} must contain exactly one unique non-empty value; got {unique!r}.")
    return unique[0]


def data_row_indices(ws, start_row, columns):
    rows = []
    for row in range(start_row, ws.max_row + 1):
        values = [normalize_cell(ws.cell(row=row, column=column).value) for column in columns]
        if any(values):
            rows.append(row)
    return rows


def lookup_mrp_description(ws, mrp_code):
    headers, _ = resolve_columns(ws, {"code": ["MRP编码"], "description": ["MRP描述"]})
    code_col = headers.get("code", 1)
    desc_col = headers.get("description", 2)
    matches = []
    for row in range(2, ws.max_row + 1):
        code = normalize_cell(ws.cell(row=row, column=code_col).value)
        if code == mrp_code:
            matches.append(normalize_cell(ws.cell(row=row, column=desc_col).value))
    matches = [item for item in matches if item]
    if len(matches) != 1:
        raise ValueError(f"MRP controller {mrp_code!r} must match exactly one description; got {matches!r}.")
    return matches[0]


def mrp_aliases(description):
    raw = normalize_cell(description)
    aliases = []
    cleaned = re.sub(r"[（(]\s*禁用\s*[)）]", "", raw).strip()
    cleaned = re.sub(r"\s*PDT.*$", "", cleaned, flags=re.IGNORECASE).strip()
    for value in [cleaned, raw.split()[0] if raw.split() else "", raw]:
        value = value.strip()
        if value and value not in aliases:
            aliases.append(value)
    return aliases


def strip_alias_from_department(user_department, aliases):
    value = normalize_cell(user_department)
    for alias in sorted(aliases, key=len, reverse=True):
        if value.startswith(alias):
            return value[len(alias):].strip()
        if value.endswith(alias):
            return value[:-len(alias)].strip()
        if alias in value:
            return value.replace(alias, "").strip()
    return value


def cost_center_candidates(user_department, aliases):
    dept_core = strip_alias_from_department(user_department, aliases)
    candidates = []
    for alias in aliases:
        if dept_core:
            candidates.append(f"{dept_core}({alias})")
            candidates.append(f"{dept_core}（{alias}）")
    candidates.append(normalize_cell(user_department))
    for alias in aliases:
        candidates.append(alias)
    return [item for index, item in enumerate(candidates) if item and item not in candidates[:index]]


def resolve_cost_center(ws, factory_code, user_department, aliases):
    columns, _ = resolve_columns(ws, COST_HEADER_ALIASES)
    missing = [key for key in COST_HEADER_ALIASES if key not in columns]
    if missing:
        raise ValueError(f"Cost center sheet is missing required columns: {missing!r}.")

    rows = []
    for row in range(2, ws.max_row + 1):
        company = normalize_cell(ws.cell(row=row, column=columns["company"]).value)
        if company != factory_code:
            continue
        rows.append({
            "row": row,
            "costCenterCode": normalize_cell(ws.cell(row=row, column=columns["costCenterCode"]).value),
            "company": company,
            "shortText": normalize_cell(ws.cell(row=row, column=columns["shortText"]).value),
        })

    candidates = cost_center_candidates(user_department, aliases)
    exact_matches = [row for row in rows if row["shortText"] in candidates]
    if len(exact_matches) == 1:
        match = exact_matches[0]
        match["searchName"] = match["shortText"]
        match["candidateNames"] = candidates
        return match

    dept_core = strip_alias_from_department(user_department, aliases)
    fuzzy_matches = [
        row for row in rows
        if row["shortText"] and dept_core and dept_core in row["shortText"]
        and any(alias in row["shortText"] for alias in aliases)
    ]
    if len(fuzzy_matches) == 1:
        match = fuzzy_matches[0]
        match["searchName"] = match["shortText"]
        match["candidateNames"] = candidates
        return match

    preview = [row["shortText"] for row in (exact_matches or fuzzy_matches or rows[:10])]
    raise ValueError(
        "Cost center lookup must resolve exactly one row. "
        f"factory={factory_code!r}, userDepartment={user_department!r}, aliases={aliases!r}, candidates={candidates!r}, preview={preview!r}"
    )


def run(input_path, user_department):
    if not user_department:
        raise ValueError("User department is required.")

    input_file = Path(input_path).expanduser().resolve()
    if not input_file.exists():
        raise FileNotFoundError(f"Excel file does not exist: {input_file}")

    wb = load_workbook(input_file, data_only=True)
    main_ws = find_sheet(wb, MAIN_SHEET_NAMES, fallback_index=0)
    mrp_ws = find_sheet(wb, MRP_SHEET_NAMES)
    cost_ws = find_sheet(wb, COST_CENTER_SHEET_NAMES)

    columns, headers = resolve_columns(main_ws, MAIN_HEADER_ALIASES)
    required = ["demandFactoryCode", "wbsCode", "projectDefinition", "mrpController"]
    missing = [key for key in required if key not in columns]
    if missing:
        raise ValueError(f"Main sheet is missing required columns: {missing!r}.")

    rows = data_row_indices(main_ws, 3, [columns[key] for key in required])
    if not rows:
        raise ValueError("Main sheet contains no data rows from row 3 onward.")

    demand_factory_code = unique_non_empty(
        [normalize_cell(main_ws.cell(row=row, column=columns["demandFactoryCode"]).value) for row in rows],
        "需求工厂代码",
    )
    wbs_code = unique_non_empty(
        [normalize_cell(main_ws.cell(row=row, column=columns["wbsCode"]).value) for row in rows],
        "WBS编码",
    )
    project_definition = unique_non_empty(
        [normalize_cell(main_ws.cell(row=row, column=columns["projectDefinition"]).value) for row in rows],
        "项目定义",
    )
    mrp_controller = unique_non_empty(
        [normalize_cell(main_ws.cell(row=row, column=columns["mrpController"]).value) for row in rows],
        "MRP控制者",
    )

    mrp_description = lookup_mrp_description(mrp_ws, mrp_controller)
    aliases = mrp_aliases(mrp_description)
    cost_center = resolve_cost_center(cost_ws, demand_factory_code, user_department, aliases)

    material_rows = []
    for row in rows:
        material_rows.append({
            "row": row,
            "materialCode": normalize_cell(main_ws.cell(row=row, column=columns.get("materialCode", 0)).value) if columns.get("materialCode") else "",
            "materialName": normalize_cell(main_ws.cell(row=row, column=columns.get("materialName", 0)).value) if columns.get("materialName") else "",
            "demandQuantity": normalize_cell(main_ws.cell(row=row, column=columns.get("demandQuantity", 0)).value) if columns.get("demandQuantity") else "",
            "unit": normalize_cell(main_ws.cell(row=row, column=columns.get("unit", 0)).value) if columns.get("unit") else "",
        })

    return {
        "ok": True,
        "inputPath": str(input_file),
        "sheetName": main_ws.title,
        "headerRow": 1,
        "instructionRow": 2,
        "dataStartRow": 3,
        "dataRows": rows,
        "headers": headers,
        "columns": columns,
        "projectDefinition": project_definition,
        "wbsCode": wbs_code,
        "demandFactoryCode": demand_factory_code,
        "mrpController": mrp_controller,
        "mrpDescription": mrp_description,
        "mrpAliases": aliases,
        "userDepartment": user_department,
        "costCenter": cost_center,
        "materialRows": material_rows,
    }


def main():
    parser = argparse.ArgumentParser(description="Validate and normalize an OA workflow 412 outbound workbook.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--user-department", required=True)
    args = parser.parse_args()
    try:
        result = run(args.input, args.user_department)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as error:
        print(json.dumps({"ok": False, "error": str(error)}, ensure_ascii=False, indent=2), file=sys.stderr)
        raise SystemExit(1)


if __name__ == "__main__":
    main()
