#!/usr/bin/env python
import argparse
import json
from datetime import date, datetime, timedelta
from pathlib import Path
import re
import sys

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing Python dependency: openpyxl. Install it with `python -m pip install openpyxl`."
    ) from exc


REQUIRED_HEADERS = {
    "demandDate": ["需求日期"],
    "projectDefinition": ["项目定义"],
    "wbsCode": ["WBS编码", "WBS项目编码"],
    "demandFactoryCode": ["需求工厂代码", "工厂代码"],
}

FACTORY_COMPANY_NAMES = {
    "1000": "北京镁伽机器人科技有限公司",
    "1010": "苏州镁伽科技有限公司",
    "1020": "深圳镁伽科技有限公司",
    "1030": "北京镁伽科技有限公司",
    "1050": "上海镁伽智能科技有限公司",
    "8040": "杭州镁伽半导体新材料有限公司",
}


def normalize_header(value):
    return re.sub(r"\s+", "", str(value or "")).strip()


def normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, date):
        return value.strftime("%Y%m%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def find_column(headers, aliases):
    normalized_aliases = {normalize_header(alias) for alias in aliases}
    for index, header in enumerate(headers, start=1):
        if normalize_header(header) in normalized_aliases:
            return index
    return None


def unique_non_empty(values, field_name):
    unique = sorted({value for value in values if value})
    if len(unique) != 1:
        raise ValueError(f"{field_name} must contain exactly one unique non-empty value; got {unique!r}.")
    return unique[0]


def data_row_indices(ws, start_row, required_columns):
    rows = []
    for row in range(start_row, ws.max_row + 1):
        values = [normalize_cell(ws.cell(row=row, column=column).value) for column in required_columns]
        if any(values):
            rows.append(row)
    return rows


def run(input_path, output_dir, days_offset):
    input_file = Path(input_path).expanduser().resolve()
    if not input_file.exists():
        raise FileNotFoundError(f"Excel file does not exist: {input_file}")

    wb = load_workbook(input_file)
    ws = wb.worksheets[0]
    headers = [ws.cell(row=1, column=column).value for column in range(1, ws.max_column + 1)]

    columns = {}
    for key, aliases in REQUIRED_HEADERS.items():
        column = find_column(headers, aliases)
        if not column:
            raise ValueError(f"Sheet1 is missing required header: {aliases[0]}")
        columns[key] = column

    rows = data_row_indices(ws, 3, columns.values())
    if not rows:
        raise ValueError("Sheet1 contains no data rows from row 3 onward.")

    project_definition = unique_non_empty(
        [normalize_cell(ws.cell(row=row, column=columns["projectDefinition"]).value) for row in rows],
        "项目定义",
    )
    wbs_code = unique_non_empty(
        [normalize_cell(ws.cell(row=row, column=columns["wbsCode"]).value) for row in rows],
        "WBS编码",
    )
    demand_factory_code = unique_non_empty(
        [normalize_cell(ws.cell(row=row, column=columns["demandFactoryCode"]).value) for row in rows],
        "需求工厂代码",
    )

    target_date = date.today() + timedelta(days=days_offset)
    target_date_text = target_date.strftime("%Y%m%d")
    original_dates = []
    for row in rows:
        cell = ws.cell(row=row, column=columns["demandDate"])
        original_dates.append(normalize_cell(cell.value))
        cell.value = target_date_text
        cell.number_format = "@"

    output_root = Path(output_dir).expanduser().resolve()
    output_root.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output_file = output_root / f"{input_file.stem}-normalized-{stamp}{input_file.suffix}"
    wb.save(output_file)

    return {
        "ok": True,
        "inputPath": str(input_file),
        "normalizedPath": str(output_file),
        "sheetName": ws.title,
        "headerRow": 1,
        "instructionRow": 2,
        "dataStartRow": 3,
        "dataRows": rows,
        "headers": [normalize_cell(item) for item in headers],
        "columns": columns,
        "targetDemandDate": target_date_text,
        "originalDemandDates": original_dates,
        "projectDefinition": project_definition,
        "wbsCode": wbs_code,
        "demandFactoryCode": demand_factory_code,
        "demandCompanyName": FACTORY_COMPANY_NAMES.get(demand_factory_code),
    }


def main():
    parser = argparse.ArgumentParser(description="Validate and normalize Sheet1 of an OA purchase request workbook.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--days-offset", type=int, default=5)
    args = parser.parse_args()
    try:
        result = run(args.input, args.output_dir, args.days_offset)
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as error:
        print(json.dumps({"ok": False, "error": str(error)}, ensure_ascii=False, indent=2), file=sys.stderr)
        raise SystemExit(1)


if __name__ == "__main__":
    main()
