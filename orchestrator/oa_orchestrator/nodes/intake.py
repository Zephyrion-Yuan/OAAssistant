"""intake node — parse the Excel into a structured BusinessInput (openpyxl),
persist it to the business_inputs SQLite table, and put it on the state.

This is the port of scripts/stock_transfer_excel.py into the Python orchestrator
(decision: Excel no longer flows through Node). Parsing logic is kept faithful:
sheet 项目需求填写界面, header row 1, data from row 3, alias-matched columns,
quantities aggregated per materialCode.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from ..config import get_settings
from ..intake_parsers import parse_inbound, parse_outbound, parse_purchase
from ..schemas import BusinessInput, MaterialPlan
from ..state import STATUS_FAILED
from .. import store

MAIN_SHEET_NAMES = ["项目需求填写界面"]

MAIN_HEADER_ALIASES = {
    "demandFactoryCode": ["需求工厂代码", "工厂代码"],
    "wbsCode": ["WBS编码", "WBS项目编码"],
    "projectDefinition": ["项目定义"],
    "materialCode": ["物料编码"],
    "materialName": ["物料名称", "物料描述"],
    "purchaseQuantity": ["需求数量", "采购数量", "数量"],
    "unit": ["基本计量单位", "单位"],
    "mrpController": ["MRP控制者", "MRP编码"],
}

REQUIRED = ["demandFactoryCode", "wbsCode", "projectDefinition", "materialCode", "purchaseQuantity"]


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, date):
        return value.strftime("%Y%m%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_decimal(value: Any, field_name: str) -> Decimal:
    text = _normalize_cell(value)
    if not text:
        raise ValueError(f"{field_name} is empty.")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{field_name} must be numeric; got {text!r}.") from exc


def _decimal_text(value: Decimal) -> str:
    normalized = value.normalize()
    if normalized == normalized.to_integral():
        return str(normalized.quantize(Decimal(1)))
    return format(normalized, "f")


def _find_sheet(workbook, names, fallback_index: Optional[int] = None):
    for name in names:
        if name in workbook.sheetnames:
            return workbook[name]
    if fallback_index is not None and fallback_index < len(workbook.worksheets):
        return workbook.worksheets[fallback_index]
    raise ValueError(f"Workbook is missing sheet: {names[0]}")


def _find_column(headers, aliases) -> Optional[int]:
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    for index, header in enumerate(headers, start=1):
        if _normalize_header(header) in normalized_aliases:
            return index
    return None


def _resolve_columns(ws):
    headers = [ws.cell(row=1, column=column).value for column in range(1, ws.max_column + 1)]
    columns: Dict[str, int] = {}
    for key, aliases in MAIN_HEADER_ALIASES.items():
        column = _find_column(headers, aliases)
        if column:
            columns[key] = column
    return columns, [_normalize_cell(item) for item in headers]


def _unique_non_empty(values: List[str], field_name: str) -> str:
    unique = sorted({value for value in values if value})
    if len(unique) != 1:
        raise ValueError(f"{field_name} must contain exactly one unique non-empty value; got {unique!r}.")
    return unique[0]


def _data_row_indices(ws, start_row: int, columns: List[int]) -> List[int]:
    rows = []
    for row in range(start_row, ws.max_row + 1):
        values = [_normalize_cell(ws.cell(row=row, column=column).value) for column in columns]
        if any(values):
            rows.append(row)
    return rows


def parse_workbook(input_path: str) -> BusinessInput:
    """Parse a purchase workbook into a BusinessInput (workflow 89 shape)."""
    input_file = Path(input_path).expanduser().resolve()
    if not input_file.exists():
        raise FileNotFoundError(f"Excel file does not exist: {input_file}")

    wb = load_workbook(input_file, data_only=True)
    ws = _find_sheet(wb, MAIN_SHEET_NAMES, fallback_index=0)
    columns, _headers = _resolve_columns(ws)
    missing = [key for key in REQUIRED if key not in columns]
    if missing:
        raise ValueError(f"Main sheet is missing required columns: {missing!r}.")

    rows = _data_row_indices(ws, 3, [columns[key] for key in REQUIRED])
    if not rows:
        raise ValueError("Main sheet contains no data rows from row 3 onward.")

    demand_factory_code = _unique_non_empty(
        [_normalize_cell(ws.cell(row=r, column=columns["demandFactoryCode"]).value) for r in rows],
        "需求工厂代码",
    )
    wbs_code = _unique_non_empty(
        [_normalize_cell(ws.cell(row=r, column=columns["wbsCode"]).value) for r in rows],
        "WBS编码",
    )
    project_definition = _unique_non_empty(
        [_normalize_cell(ws.cell(row=r, column=columns["projectDefinition"]).value) for r in rows],
        "项目定义",
    )

    quantity_totals: Dict[str, Decimal] = {}
    material_names: Dict[str, str] = {}
    material_units: Dict[str, str] = {}
    for r in rows:
        material_code = _normalize_cell(ws.cell(row=r, column=columns["materialCode"]).value)
        if not material_code:
            raise ValueError(f"Row {r} is missing 物料编码.")
        quantity = _parse_decimal(ws.cell(row=r, column=columns["purchaseQuantity"]).value, f"Row {r} 数量")
        material_name = _normalize_cell(ws.cell(row=r, column=columns.get("materialName", 0)).value) if columns.get("materialName") else ""
        unit = _normalize_cell(ws.cell(row=r, column=columns.get("unit", 0)).value) if columns.get("unit") else ""
        quantity_totals[material_code] = quantity_totals.get(material_code, Decimal(0)) + quantity
        material_names.setdefault(material_code, material_name)
        material_units.setdefault(material_code, unit)

    material_plans = [
        MaterialPlan(
            materialCode=code,
            materialName=material_names.get(code, ""),
            quantity=_decimal_text(qty),
            unit=material_units.get(code, ""),
        )
        for code, qty in sorted(quantity_totals.items())
    ]
    mrp_controller = (
        _normalize_cell(ws.cell(row=rows[0], column=columns.get("mrpController", 0)).value)
        if columns.get("mrpController") else ""
    )

    return BusinessInput(
        projectDefinition=project_definition,
        wbsCode=wbs_code,
        demandFactoryCode=demand_factory_code,
        mrpController=mrp_controller,
        materialPlans=material_plans,
        quantityByMaterialCode={code: _decimal_text(qty) for code, qty in sorted(quantity_totals.items())},
        sourceFile=str(input_file),
    )


def _material_plans_from_rows(rows: List[Dict[str, Any]], quantity_key: str) -> List[MaterialPlan]:
    """Aggregate per-workflow Excel `materialRows` into MaterialPlans so the
    workflow-agnostic pdm_enrich node can validate the codes."""
    quantity_totals: Dict[str, Decimal] = {}
    names: Dict[str, str] = {}
    units: Dict[str, str] = {}
    for row in rows or []:
        code = _normalize_cell(row.get("materialCode"))
        if not code:
            continue
        raw_qty = row.get(quantity_key)
        try:
            qty = _parse_decimal(raw_qty, "数量") if _normalize_cell(raw_qty) else Decimal(0)
        except ValueError:
            qty = Decimal(0)
        quantity_totals[code] = quantity_totals.get(code, Decimal(0)) + qty
        names.setdefault(code, _normalize_cell(row.get("materialName")))
        units.setdefault(code, _normalize_cell(row.get("unit")))
    return [
        MaterialPlan(materialCode=code, materialName=names.get(code, ""),
                     quantity=_decimal_text(qty), unit=units.get(code, ""))
        for code, qty in sorted(quantity_totals.items())
    ]


def _business_from_outbound(structured: Dict[str, Any]) -> BusinessInput:
    plans = _material_plans_from_rows(structured.get("materialRows", []), "demandQuantity")
    return BusinessInput(
        projectDefinition=structured.get("projectDefinition"),
        wbsCode=structured.get("wbsCode"),
        demandFactoryCode=structured.get("demandFactoryCode"),
        mrpController=structured.get("mrpController"),
        materialPlans=plans,
        quantityByMaterialCode={p.materialCode: p.quantity for p in plans},
        sourceFile=structured.get("sourceFile"),
        structured=structured,
    )


def _business_from_inbound(structured: Dict[str, Any]) -> BusinessInput:
    plans = _material_plans_from_rows(structured.get("materialRows", []), "purchaseQuantity")
    return BusinessInput(
        projectDefinition=structured.get("projectDefinition"),
        wbsCode=structured.get("wbsCode"),
        demandFactoryCode=structured.get("demandFactoryCode"),
        mrpController=structured.get("mrpController"),
        materialPlans=plans,
        quantityByMaterialCode=dict(structured.get("quantityByMaterialCode", {})),
        sourceFile=structured.get("sourceFile"),
        structured=structured,
    )


def _business_from_purchase(structured: Dict[str, Any]) -> BusinessInput:
    # 458 is attachment-driven: no material plans to PDM-validate.
    return BusinessInput(
        projectDefinition=structured.get("projectDefinition"),
        wbsCode=structured.get("wbsCode"),
        demandFactoryCode=structured.get("demandFactoryCode"),
        materialPlans=[],
        sourceFile=structured.get("sourceFile"),
        structured=structured,
    )


def _parse_by_workflow(workflow_id: str, excel_path: str, state: Dict[str, Any]) -> BusinessInput:
    """Select the per-workflow intake parser. 89 keeps using parse_workbook."""
    if workflow_id == "412":
        user_department = (state.get("profile") or {}).get("department") or state.get("user_department") or ""
        return _business_from_outbound(parse_outbound(excel_path, user_department))
    if workflow_id == "414":
        return _business_from_inbound(parse_inbound(excel_path))
    if workflow_id == "458":
        return _business_from_purchase(parse_purchase(excel_path))
    return parse_workbook(excel_path)


def intake_node(state: Dict[str, Any]) -> Dict[str, Any]:
    settings = get_settings()
    history = list(state.get("history", []))
    excel_path = state.get("excel_path")
    workflow_id = state.get("workflow_id", "89")
    # Allow resuming a thread whose business_input was parsed previously.
    if state.get("business_input"):
        return {}
    if not excel_path:
        existing = store.get_business_input(settings.store_path, state.get("thread_id", ""))
        if existing:
            history.append({"node": "intake", "ok": True, "source": "store"})
            return {"business_input": existing.model_dump(), "history": history}
        history.append({"node": "intake", "ok": False, "error": "No excel_path and no stored input."})
        return {"status": STATUS_FAILED, "history": history,
                "result": {"ok": False, "error": "No excel_path provided and no stored business input."}}
    try:
        business = _parse_by_workflow(workflow_id, excel_path, state)
    except Exception as exc:  # noqa: BLE001
        history.append({"node": "intake", "ok": False, "error": str(exc)})
        return {"status": STATUS_FAILED, "history": history,
                "result": {"ok": False, "error": f"intake failed: {exc}"}}

    store.save_business_input(settings.store_path, state.get("thread_id", ""), business, business.sourceFile)
    history.append({"node": "intake", "ok": True, "workflow": workflow_id,
                    "materials": len(business.materialPlans), "source": "excel"})
    return {"business_input": business.model_dump(), "history": history}
