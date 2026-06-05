"""Per-workflow Excel parsers (openpyxl ports of the scripts/*_excel.py helpers).

Each parser returns the *structured dict* the matching Node endpoint expects —
the same shape the scripts/<wf>_excel.py CLI helpers print to stdout. The intake
node selects the parser by state['workflow_id'] (89 keeps using parse_workbook;
412/414/458 use the parsers here).

  parse_outbound(path, user_department) -> dict   # workflow 412 (物资出库)
  parse_inbound(path)                   -> dict   # workflow 414 (物资入库)
  parse_purchase(path, output_dir, ..)  -> dict   # workflow 458 (采购申请)

These are faithful ports of:
  scripts/outbound_excel.py, scripts/inbound_excel.py, scripts/purchase_excel.py
Parsing logic (sheet names, header aliases, header row 1 / data from row 3,
cost-center resolution, attachment normalization) is kept 1:1.
"""
from __future__ import annotations

import re
import tempfile
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


# --------------------------------------------------------------------------- #
# Shared helpers (mirror the *_excel.py scripts)
# --------------------------------------------------------------------------- #
def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d")
    if isinstance(value, date):
        return value.strftime("%Y%m%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_decimal(value: Any, field_name: str) -> Decimal:
    text = _normalize_cell(value)
    if not text:
        raise ValueError(f"{field_name} is empty.")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{field_name} must be numeric; got {text!r}.") from exc


def _decimal_text(value: Decimal) -> str:
    normalized = value.normalize()
    if normalized == normalized.to_integral():
        return str(normalized.quantize(Decimal(1)))
    return format(normalized, "f")


def _find_sheet(workbook, names, fallback_index: Optional[int] = None):
    for name in names:
        if name in workbook.sheetnames:
            return workbook[name]
    if fallback_index is not None and fallback_index < len(workbook.worksheets):
        return workbook.worksheets[fallback_index]
    raise ValueError(f"Workbook is missing sheet: {names[0]}")


def _find_column(headers, aliases) -> Optional[int]:
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    for index, header in enumerate(headers, start=1):
        if _normalize_header(header) in normalized_aliases:
            return index
    return None


def _resolve_columns(ws, aliases_by_key):
    headers = [ws.cell(row=1, column=column).value for column in range(1, ws.max_column + 1)]
    columns: Dict[str, int] = {}
    for key, aliases in aliases_by_key.items():
        column = _find_column(headers, aliases)
        if column:
            columns[key] = column
    return columns, [_normalize_cell(item) for item in headers]


def _unique_non_empty(values: List[str], field_name: str) -> str:
    unique = sorted({value for value in values if value})
    if len(unique) != 1:
        raise ValueError(f"{field_name} must contain exactly one unique non-empty value; got {unique!r}.")
    return unique[0]


def _data_row_indices(ws, start_row: int, columns: List[int]) -> List[int]:
    rows = []
    for row in range(start_row, ws.max_row + 1):
        values = [_normalize_cell(ws.cell(row=row, column=column).value) for column in columns]
        if any(values):
            rows.append(row)
    return rows


FACTORY_COMPANY_NAMES = {
    "1000": "北京镁伽机器人科技有限公司",
    "1010": "苏州镁伽科技有限公司",
    "1020": "深圳镁伽科技有限公司",
    "1030": "北京镁伽科技有限公司",
    "1050": "上海镁伽智能科技有限公司",
    "8040": "杭州镁伽半导体新材料有限公司",
}


# --------------------------------------------------------------------------- #
# Workflow 412 — outbound (物资出库). Port of scripts/outbound_excel.py
# --------------------------------------------------------------------------- #
_OUTBOUND_MAIN_SHEETS = ["项目需求填写界面"]
_OUTBOUND_MRP_SHEETS = ["MRP控制者"]
_OUTBOUND_COST_SHEETS = ["成本中心"]

_OUTBOUND_MAIN_ALIASES = {
    "demandFactoryCode": ["需求工厂代码", "工厂代码"],
    "wbsCode": ["WBS编码", "WBS项目编码"],
    "projectDefinition": ["项目定义"],
    "mrpController": ["MRP控制者", "MRP编码"],
    "materialCode": ["物料编码"],
    "materialName": ["物料名称"],
    "demandQuantity": ["需求数量"],
    "unit": ["基本计量单位", "采购单位", "单位"],
}

_OUTBOUND_COST_ALIASES = {
    "costCenterCode": ["成本中心"],
    "company": ["公司", "公司代码"],
    "shortText": ["短文本", "成本中心名称"],
}


def _lookup_mrp_description(ws, mrp_code: str) -> str:
    headers, _ = _resolve_columns(ws, {"code": ["MRP编码"], "description": ["MRP描述"]})
    code_col = headers.get("code", 1)
    desc_col = headers.get("description", 2)
    matches = []
    for row in range(2, ws.max_row + 1):
        code = _normalize_cell(ws.cell(row=row, column=code_col).value)
        if code == mrp_code:
            matches.append(_normalize_cell(ws.cell(row=row, column=desc_col).value))
    matches = [item for item in matches if item]
    if len(matches) != 1:
        raise ValueError(f"MRP controller {mrp_code!r} must match exactly one description; got {matches!r}.")
    return matches[0]


def _mrp_aliases(description: str) -> List[str]:
    raw = _normalize_cell(description)
    aliases: List[str] = []
    cleaned = re.sub(r"[（(]\s*禁用\s*[)）]", "", raw).strip()
    cleaned = re.sub(r"\s*PDT.*$", "", cleaned, flags=re.IGNORECASE).strip()
    for value in [cleaned, raw.split()[0] if raw.split() else "", raw]:
        value = value.strip()
        if value and value not in aliases:
            aliases.append(value)
    return aliases


def _strip_alias_from_department(user_department: str, aliases: List[str]) -> str:
    value = _normalize_cell(user_department)
    for alias in sorted(aliases, key=len, reverse=True):
        if value.startswith(alias):
            return value[len(alias):].strip()
        if value.endswith(alias):
            return value[:-len(alias)].strip()
        if alias in value:
            return value.replace(alias, "").strip()
    return value


def _cost_center_candidates(user_department: str, aliases: List[str]) -> List[str]:
    dept_core = _strip_alias_from_department(user_department, aliases)
    candidates: List[str] = []
    for alias in aliases:
        if dept_core:
            candidates.append(f"{dept_core}({alias})")
            candidates.append(f"{dept_core}（{alias}）")
    candidates.append(_normalize_cell(user_department))
    for alias in aliases:
        candidates.append(alias)
    return [item for index, item in enumerate(candidates) if item and item not in candidates[:index]]


def _resolve_cost_center(ws, factory_code: str, user_department: str, aliases: List[str]) -> Dict[str, Any]:
    columns, _ = _resolve_columns(ws, _OUTBOUND_COST_ALIASES)
    missing = [key for key in _OUTBOUND_COST_ALIASES if key not in columns]
    if missing:
        raise ValueError(f"Cost center sheet is missing required columns: {missing!r}.")

    rows = []
    for row in range(2, ws.max_row + 1):
        company = _normalize_cell(ws.cell(row=row, column=columns["company"]).value)
        if company != factory_code:
            continue
        rows.append({
            "row": row,
            "costCenterCode": _normalize_cell(ws.cell(row=row, column=columns["costCenterCode"]).value),
            "company": company,
            "shortText": _normalize_cell(ws.cell(row=row, column=columns["shortText"]).value),
        })

    candidates = _cost_center_candidates(user_department, aliases)
    exact_matches = [row for row in rows if row["shortText"] in candidates]
    if len(exact_matches) == 1:
        match = exact_matches[0]
        match["searchName"] = match["shortText"]
        match["candidateNames"] = candidates
        return match

    dept_core = _strip_alias_from_department(user_department, aliases)
    fuzzy_matches = [
        row for row in rows
        if row["shortText"] and dept_core and dept_core in row["shortText"]
        and any(alias in row["shortText"] for alias in aliases)
    ]
    if len(fuzzy_matches) == 1:
        match = fuzzy_matches[0]
        match["searchName"] = match["shortText"]
        match["candidateNames"] = candidates
        return match

    preview = [row["shortText"] for row in (exact_matches or fuzzy_matches or rows[:10])]
    raise ValueError(
        "Cost center lookup must resolve exactly one row. "
        f"factory={factory_code!r}, userDepartment={user_department!r}, "
        f"aliases={aliases!r}, candidates={candidates!r}, preview={preview!r}"
    )


def parse_outbound(input_path: str, user_department: str) -> Dict[str, Any]:
    """Parse a workflow 412 outbound workbook -> structured dict (Node shape)."""
    if not user_department:
        raise ValueError("User department is required.")

    input_file = Path(input_path).expanduser().resolve()
    if not input_file.exists():
        raise FileNotFoundError(f"Excel file does not exist: {input_file}")

    wb = load_workbook(input_file, data_only=True)
    main_ws = _find_sheet(wb, _OUTBOUND_MAIN_SHEETS, fallback_index=0)
    mrp_ws = _find_sheet(wb, _OUTBOUND_MRP_SHEETS)
    cost_ws = _find_sheet(wb, _OUTBOUND_COST_SHEETS)

    columns, headers = _resolve_columns(main_ws, _OUTBOUND_MAIN_ALIASES)
    required = ["demandFactoryCode", "wbsCode", "projectDefinition", "mrpController"]
    missing = [key for key in required if key not in columns]
    if missing:
        raise ValueError(f"Main sheet is missing required columns: {missing!r}.")

    rows = _data_row_indices(main_ws, 3, [columns[key] for key in required])
    if not rows:
        raise ValueError("Main sheet contains no data rows from row 3 onward.")

    demand_factory_code = _unique_non_empty(
        [_normalize_cell(main_ws.cell(row=r, column=columns["demandFactoryCode"]).value) for r in rows],
        "需求工厂代码",
    )
    wbs_code = _unique_non_empty(
        [_normalize_cell(main_ws.cell(row=r, column=columns["wbsCode"]).value) for r in rows],
        "WBS编码",
    )
    project_definition = _unique_non_empty(
        [_normalize_cell(main_ws.cell(row=r, column=columns["projectDefinition"]).value) for r in rows],
        "项目定义",
    )
    mrp_controller = _unique_non_empty(
        [_normalize_cell(main_ws.cell(row=r, column=columns["mrpController"]).value) for r in rows],
        "MRP控制者",
    )

    mrp_description = _lookup_mrp_description(mrp_ws, mrp_controller)
    aliases = _mrp_aliases(mrp_description)
    cost_center = _resolve_cost_center(cost_ws, demand_factory_code, user_department, aliases)

    material_rows = []
    for r in rows:
        material_rows.append({
            "row": r,
            "materialCode": _normalize_cell(main_ws.cell(row=r, column=columns.get("materialCode", 0)).value) if columns.get("materialCode") else "",
            "materialName": _normalize_cell(main_ws.cell(row=r, column=columns.get("materialName", 0)).value) if columns.get("materialName") else "",
            "demandQuantity": _normalize_cell(main_ws.cell(row=r, column=columns.get("demandQuantity", 0)).value) if columns.get("demandQuantity") else "",
            "unit": _normalize_cell(main_ws.cell(row=r, column=columns.get("unit", 0)).value) if columns.get("unit") else "",
        })

    return {
        "ok": True,
        "inputPath": str(input_file),
        "sheetName": main_ws.title,
        "headerRow": 1,
        "instructionRow": 2,
        "dataStartRow": 3,
        "dataRows": rows,
        "headers": headers,
        "columns": columns,
        "projectDefinition": project_definition,
        "wbsCode": wbs_code,
        "demandFactoryCode": demand_factory_code,
        "mrpController": mrp_controller,
        "mrpDescription": mrp_description,
        "mrpAliases": aliases,
        "userDepartment": user_department,
        "costCenter": cost_center,
        "materialRows": material_rows,
        "sourceFile": str(input_file),
    }


# --------------------------------------------------------------------------- #
# Workflow 414 — inbound (物资入库). Port of scripts/inbound_excel.py
# --------------------------------------------------------------------------- #
_INBOUND_MAIN_SHEETS = ["项目需求填写界面"]

_INBOUND_MAIN_ALIASES = {
    "demandFactoryCode": ["需求工厂代码", "工厂代码"],
    "wbsCode": ["WBS编码", "WBS项目编码"],
    "projectDefinition": ["项目定义"],
    "materialCode": ["物料编码"],
    "materialName": ["物料名称"],
    "purchaseQuantity": ["需求数量", "采购数量", "数量"],
    "unit": ["基本计量单位", "采购单位", "单位"],
    "mrpController": ["MRP控制者", "MRP编码"],
}


def parse_inbound(input_path: str) -> Dict[str, Any]:
    """Parse a workflow 414 inbound workbook -> structured dict (Node shape)."""
    input_file = Path(input_path).expanduser().resolve()
    if not input_file.exists():
        raise FileNotFoundError(f"Excel file does not exist: {input_file}")

    wb = load_workbook(input_file, data_only=True)
    main_ws = _find_sheet(wb, _INBOUND_MAIN_SHEETS, fallback_index=0)
    columns, headers = _resolve_columns(main_ws, _INBOUND_MAIN_ALIASES)
    required = ["demandFactoryCode", "wbsCode", "projectDefinition", "materialCode", "purchaseQuantity"]
    missing = [key for key in required if key not in columns]
    if missing:
        raise ValueError(f"Main sheet is missing required columns: {missing!r}.")

    rows = _data_row_indices(main_ws, 3, [columns[key] for key in required])
    if not rows:
        raise ValueError("Main sheet contains no data rows from row 3 onward.")

    demand_factory_code = _unique_non_empty(
        [_normalize_cell(main_ws.cell(row=r, column=columns["demandFactoryCode"]).value) for r in rows],
        "需求工厂代码",
    )
    wbs_code = _unique_non_empty(
        [_normalize_cell(main_ws.cell(row=r, column=columns["wbsCode"]).value) for r in rows],
        "WBS编码",
    )
    project_definition = _unique_non_empty(
        [_normalize_cell(main_ws.cell(row=r, column=columns["projectDefinition"]).value) for r in rows],
        "项目定义",
    )

    material_rows = []
    quantity_totals: Dict[str, Decimal] = {}
    for r in rows:
        material_code = _normalize_cell(main_ws.cell(row=r, column=columns["materialCode"]).value)
        if not material_code:
            raise ValueError(f"Row {r} is missing 物料编码.")
        quantity = _parse_decimal(main_ws.cell(row=r, column=columns["purchaseQuantity"]).value, f"Row {r} 需求数量")
        quantity_totals[material_code] = quantity_totals.get(material_code, Decimal(0)) + quantity
        material_rows.append({
            "row": r,
            "materialCode": material_code,
            "materialName": _normalize_cell(main_ws.cell(row=r, column=columns.get("materialName", 0)).value) if columns.get("materialName") else "",
            "purchaseQuantity": _decimal_text(quantity),
            "unit": _normalize_cell(main_ws.cell(row=r, column=columns.get("unit", 0)).value) if columns.get("unit") else "",
        })

    return {
        "ok": True,
        "inputPath": str(input_file),
        "sheetName": main_ws.title,
        "headerRow": 1,
        "instructionRow": 2,
        "dataStartRow": 3,
        "dataRows": rows,
        "headers": headers,
        "columns": columns,
        "projectDefinition": project_definition,
        "wbsCode": wbs_code,
        "demandFactoryCode": demand_factory_code,
        "mrpController": _normalize_cell(main_ws.cell(row=rows[0], column=columns.get("mrpController", 0)).value) if columns.get("mrpController") else "",
        "materialRows": material_rows,
        "quantityByMaterialCode": {
            material_code: _decimal_text(quantity)
            for material_code, quantity in sorted(quantity_totals.items())
        },
        "sourceFile": str(input_file),
    }


# --------------------------------------------------------------------------- #
# Workflow 458 — purchase (采购申请). Port of scripts/purchase_excel.py
# --------------------------------------------------------------------------- #
_PURCHASE_REQUIRED_HEADERS = {
    "demandDate": ["需求日期"],
    "projectDefinition": ["项目定义"],
    "wbsCode": ["WBS编码", "WBS项目编码"],
    "demandFactoryCode": ["需求工厂代码", "工厂代码"],
}


def parse_purchase(input_path: str, output_dir: Optional[str] = None, days_offset: int = 5) -> Dict[str, Any]:
    """Parse + normalize a workflow 458 purchase workbook -> structured dict.

    Writes a normalized attachment (with the demand date rewritten to today +
    days_offset) into output_dir (defaults to a temp dir) and reports its path
    as `normalizedPath`, which the Node /api/oa/purchase endpoint uploads.
    """
    input_file = Path(input_path).expanduser().resolve()
    if not input_file.exists():
        raise FileNotFoundError(f"Excel file does not exist: {input_file}")

    wb = load_workbook(input_file)
    ws = wb.worksheets[0]
    headers = [ws.cell(row=1, column=column).value for column in range(1, ws.max_column + 1)]

    columns: Dict[str, int] = {}
    for key, aliases in _PURCHASE_REQUIRED_HEADERS.items():
        column = _find_column(headers, aliases)
        if not column:
            raise ValueError(f"Sheet1 is missing required header: {aliases[0]}")
        columns[key] = column

    rows = _data_row_indices(ws, 3, list(columns.values()))
    if not rows:
        raise ValueError("Sheet1 contains no data rows from row 3 onward.")

    project_definition = _unique_non_empty(
        [_normalize_cell(ws.cell(row=r, column=columns["projectDefinition"]).value) for r in rows],
        "项目定义",
    )
    wbs_code = _unique_non_empty(
        [_normalize_cell(ws.cell(row=r, column=columns["wbsCode"]).value) for r in rows],
        "WBS编码",
    )
    demand_factory_code = _unique_non_empty(
        [_normalize_cell(ws.cell(row=r, column=columns["demandFactoryCode"]).value) for r in rows],
        "需求工厂代码",
    )

    target_date = date.today() + timedelta(days=days_offset)
    target_date_text = target_date.strftime("%Y%m%d")
    original_dates = []
    for r in rows:
        cell = ws.cell(row=r, column=columns["demandDate"])
        original_dates.append(_normalize_cell(cell.value))
        cell.value = target_date_text
        cell.number_format = "@"

    output_root = Path(output_dir).expanduser().resolve() if output_dir else Path(tempfile.mkdtemp())
    output_root.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output_file = output_root / f"{input_file.stem}-normalized-{stamp}{input_file.suffix}"
    wb.save(output_file)

    return {
        "ok": True,
        "inputPath": str(input_file),
        "normalizedPath": str(output_file),
        "sheetName": ws.title,
        "headerRow": 1,
        "instructionRow": 2,
        "dataStartRow": 3,
        "dataRows": rows,
        "headers": [_normalize_cell(item) for item in headers],
        "columns": columns,
        "targetDemandDate": target_date_text,
        "originalDemandDates": original_dates,
        "projectDefinition": project_definition,
        "wbsCode": wbs_code,
        "demandFactoryCode": demand_factory_code,
        "demandCompanyName": FACTORY_COMPANY_NAMES.get(demand_factory_code),
        "sourceFile": str(input_file),
    }
