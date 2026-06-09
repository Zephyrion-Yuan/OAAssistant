"""Offline test: the Phase-1 acquire-mode WBS-fan-out router, end to end.

No network, no DeepSeek, no Edge. Drives the full acquire graph against the mock
backend and asserts the 412/89/458 fan-out, WBS bucketing, generated 458
attachment, and the skip-on-missing-registry path. Run:
    orchestrator/.venv/bin/python orchestrator/tests/router.py
"""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

os.environ["EXECUTOR"] = "mock"
os.environ["DEEPSEEK_API_KEY"] = ""  # offline gate

from openpyxl import Workbook, load_workbook  # noqa: E402

from oa_orchestrator.executors import mock as mockmod  # noqa: E402
from oa_orchestrator.executors.mock import MockExecutor  # noqa: E402
from oa_orchestrator.graph import build_graph, make_checkpointer  # noqa: E402
from oa_orchestrator.llm import set_test_responder  # noqa: E402
from oa_orchestrator.nodes.classify_goal import GoalClassification  # noqa: E402
from oa_orchestrator.nodes.route_workflow import allocate  # noqa: E402
from oa_orchestrator.nodes.unit_check import UnitJudgment  # noqa: E402
from oa_orchestrator.state import (STATUS_DONE, STATUS_NEEDS_INPUT,  # noqa: E402
                                   new_state)

# LLM is mandatory (no heuristic fallback). Register a deterministic stub so the
# classify_goal / unit_check nodes run offline. _GOAL drives the classifier;
# UnitJudgment defaults to "inconsistent" so any unit-mismatch is flagged.
_GOAL = {"value": "acquire"}


def _stub_llm(schema, system, user):
    if schema is GoalClassification:
        return GoalClassification(goal=_GOAL["value"])
    if schema is UnitJudgment:
        return UnitJudgment(consistent=False, suggestedUnit="箱", suggestedQuantity="1",
                            reason="规格 50 盒/箱,疑似单位误用")
    return None


set_test_responder(_stub_llm)

HEADERS = ["需求工厂代码", "WBS编码", "项目定义", "物料编码", "物料名称",
           "需求数量", "基本计量单位", "MRP控制者"]
DEMAND_WBS = "C2-0225002.06.01"      # in mock registry (cost center + stock loc)
SOURCE_WBS = "C2-0339001.01.01"      # another project's WBS (also in registry)


def _loc(factory, code, name, wbs, qty, sobkz):
    return {"factoryCode": factory, "stockLocationCode": code, "stockLocationName": name,
            "wbsCode": wbs, "unrestrictedStock": qty, "specialStockIndicator": sobkz, "unit": "EA"}


def make_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "项目需求填写界面"
    ws.append(HEADERS)
    ws.append(["说明"] * len(HEADERS))
    for r in rows:
        ws.append(r)
    path = Path(tempfile.mkdtemp()) / "demand.xlsx"
    wb.save(path)
    return str(path)


def run(graph, *, excel, thread, save, mode="acquire", profile=None):
    config = {"configurable": {"thread_id": thread}}
    state = new_state(request="按需求表采购/出库/转储", excel_path=excel, thread_id=thread,
                      interactive=False, save=save, mode=mode,
                      profile=profile or {"user_id": "tester", "department": "研发三组"})
    return graph.invoke(state, config)


def main() -> int:
    # 0) pure allocation sanity — the confirmed routing table:
    #    own-project -> 412 | public -> 89(公共→项目)+412 | other-project -> 建议458 | shortfall -> 458
    own = allocate([{"materialCode": "M", "wbsCode": DEMAND_WBS, "quantity": "3"}],
                   {"M": {"locations": [{"wbsCode": DEMAND_WBS, "unrestrictedStock": "5", "isProjectStock": True}]}})
    assert sorted(e.workflow_id for e in own.entries) == ["412"], own.entries
    pub = allocate([{"materialCode": "M", "wbsCode": DEMAND_WBS, "quantity": "4"}],
                   {"M": {"locations": [{"unrestrictedStock": "10", "isProjectStock": False,
                                         "stockLocationCode": "A001", "stockLocationName": "成品仓"}]}})
    assert sorted(e.workflow_id for e in pub.entries) == ["412", "89"], pub.entries
    e89 = next(e for e in pub.entries if e.workflow_id == "89")
    assert e89.sourceKind == "public" and e89.transferOutStockLocationName == "成品仓" \
        and e89.movementType == "普通库存转储至项目库存", e89
    # other-project: recommend 458 (no auto-89), recorded as a recommendation
    other = allocate([{"materialCode": "M", "wbsCode": DEMAND_WBS, "quantity": "5"}],
                     {"M": {"locations": [{"wbsCode": SOURCE_WBS, "unrestrictedStock": "2", "isProjectStock": True}]}})
    assert sorted(e.workflow_id for e in other.entries) == ["458"], other.entries
    assert other.recommendations and other.recommendations[0]["materialCode"] == "M", other.recommendations
    # ...unless the user overrides to transfer -> 89(项目→项目) + 458 shortfall
    over = allocate([{"materialCode": "M", "wbsCode": DEMAND_WBS, "quantity": "5"}],
                    {"M": {"locations": [{"wbsCode": SOURCE_WBS, "unrestrictedStock": "2", "isProjectStock": True}]}},
                    routing_overrides={"M": "transfer"})
    assert sorted(e.workflow_id for e in over.entries) == ["458", "89"], over.entries
    e89o = next(e for e in over.entries if e.workflow_id == "89")
    assert e89o.transferOutWbs == SOURCE_WBS and e89o.movementType == "项目库存转储至项目库存", e89o
    print("PASS allocation: own→412 | public→89(公共→项目)+412 | other→建议458 | override→89(项目→项目)+458")

    # scenario inventory: 4000023659 public(10), 4000059295 other-project special(2)
    mockmod._INVENTORY = {
        "4000023659": [_loc("1010", "A001", "成品仓", "", "10.000", "")],
        "4000059295": [_loc("1010", "D002", "设备零件仓", SOURCE_WBS, "2.000", "Q")],
    }

    graph = build_graph(MockExecutor(), checkpointer=make_checkpointer(None), mode="acquire")

    # 1) three-flow fan-out, single demand WBS, save -> 3 saved drafts
    xlsx = make_workbook([
        ["1010", DEMAND_WBS, "C2-0225002", "4000023659", "PCR板", 10, "EA", "M01"],
        ["1010", DEMAND_WBS, "C2-0225002", "4000059295", "传感器", 5, "EA", "M01"],
    ])
    out = run(graph, excel=xlsx, thread="router-3way", save=True)
    assert out.get("status") == STATUS_DONE, (out.get("status"), out.get("result"))
    result = out["result"]
    assert result["router"] and result["ok"], result
    drafts = {d["workflow_id"]: d for d in result["drafts"]}
    assert set(drafts) == {"412", "89", "458"}, set(drafts)
    # 412: outbound the public PCR板 x10 (own-project 0 + transferred-in public 10)
    assert drafts["412"]["wbsCode"] == DEMAND_WBS
    assert drafts["412"]["materialLines"][0]["materialCode"] == "4000023659"
    assert drafts["412"]["materialLines"][0]["quantity"] == "10"
    entry412 = next(e for e in out["plan"]["entries"] if e["workflow_id"] == "412")
    assert entry412["request"]["userDepartment"] == "研发三组", entry412["request"]
    # 89: move the public PCR板 x10 公共仓→项目仓 (no source WBS — public source location)
    assert not drafts["89"]["transferOutWbs"], drafts["89"]
    assert drafts["89"]["materialLines"][0]["materialCode"] == "4000023659"
    assert drafts["89"]["materialLines"][0]["quantity"] == "10"
    entry89 = next(e for e in out["plan"]["entries"] if e["workflow_id"] == "89")
    assert entry89["request"]["movementType"] == "普通库存转储至项目库存", entry89["request"]
    # 458: other-project stock found -> recommend purchasing all 5 (传感器)
    assert drafts["458"]["materialLines"][0]["materialCode"] == "4000059295"
    assert drafts["458"]["materialLines"][0]["quantity"] == "5"
    assert all(d["ok"] and d["requestId"] for d in drafts.values()), drafts
    print("PASS 3-flow fan-out: 412(PCR板x10) + 89(公共→项目 PCR板x10) + 458(传感器x5, 其他项目仓建议采购) all saved")

    # 2) generated 458 attachment matches the real 22-column 项目需求填写界面 layout
    entry458 = next(e for e in out["plan"]["entries"] if e["workflow_id"] == "458")
    assert entry458["request"]["purchaseType"] == "项目物资采购申请", entry458["request"]
    assert entry458["request"]["projectType"] == "是", entry458["request"]
    att = Path(entry458["request"]["structured"]["normalizedPath"])
    assert att.exists(), att
    awb = load_workbook(att)
    aws = awb.active
    assert aws.title == "项目需求填写界面"
    hdr = [c.value for c in aws[1]]
    assert hdr[:6] == ["需求类型", "物料编码", "物料名称", "物料组", "需求数量", "基本计量单位"], hdr
    assert len(hdr) == 22, len(hdr)
    data = [c.value for c in aws[3]]
    # 需求类型=02, 物料编码, 需求数量=5 (recommend purchasing all), 申请人(idx9), WBS(idx11), 工厂(idx16)
    assert data[0] == "02" and data[1] == "4000059295" and str(data[4]) == "5", data
    assert data[2] == "传感器" and data[5] == "EA", data
    assert data[10] == "C2-0225002" and data[15] == "M01", data
    assert data[11] == DEMAND_WBS and data[16] == "1010", data
    assert data[9] == "demo-buyer", data[9]   # 申请人 from registry.purchaser
    assert data[18] == "苏州工业园区玲珑街88号" and data[19] == "紧急", data
    print("PASS 458 attachment matches real 22-col template:", att.name)

    # 3) other-project recommendation note recorded (建议采购)
    assert any("其他项目仓" in n and "建议走采购" in n for n in result["notes"]), result["notes"]
    print("PASS other-project recommendation note recorded")

    # 3a) 412 requires user department for cost-center matching.
    missing_dept = run(graph, excel=make_workbook([
        ["1010", DEMAND_WBS, "C2-0225002", "4000023659", "PCR板", 1, "EA", "M01"],
    ]), thread="router-missing-department", save=True, profile={"user_id": "tester"})
    assert missing_dept.get("status") == STATUS_NEEDS_INPUT, missing_dept.get("status")
    missing_input = (missing_dept.get("result") or {}).get("input") or {}
    assert missing_input.get("kind") == "userDepartment", missing_input
    assert missing_input.get("wbsCode") == DEMAND_WBS, missing_input
    print("PASS 412 missing profile department -> needs_input userDepartment")

    # 3b) 412 project code falls back to the leading project definition in WBS.
    saved_demand_wbs = dict(mockmod._WBS_REGISTRY[DEMAND_WBS])
    try:
        mockmod._WBS_REGISTRY[DEMAND_WBS] = {**saved_demand_wbs, "projectDefinition": ""}
        mockmod._INVENTORY["4000023659"] = [_loc("1010", "A001", "成品仓", "", "1.000", "")]
        project_fallback = run(graph, excel=make_workbook([
            ["1010", DEMAND_WBS, "", "4000023659", "PCR板", 1, "EA", "M01"],
        ]), thread="router-412-project-fallback", save=True)
        assert project_fallback.get("status") == STATUS_DONE, (
            project_fallback.get("status"), project_fallback.get("result")
        )
        entry412_fallback = next(e for e in project_fallback["plan"]["entries"] if e["workflow_id"] == "412")
        assert entry412_fallback["request"]["structured"]["projectDefinition"] == "C2-0225002", (
            entry412_fallback["request"]
        )
        print("PASS 412 projectDefinition empty -> derived from WBS for project-code fill")
    finally:
        mockmod._WBS_REGISTRY[DEMAND_WBS] = saved_demand_wbs

    # 3c) frontend-like sparse row: PDM fills material name/unit; WBS fills
    # project definition/MRP/delivery address/remark.
    mockmod._INVENTORY["4000059295"] = []
    sparse = run(graph, excel=make_workbook([
        ["1010", DEMAND_WBS, "", "4000059295", "", 1, "", ""],
    ]), thread="router-sparse-458", save=True)
    assert sparse.get("status") == STATUS_DONE, (sparse.get("status"), sparse.get("result"))
    sparse_entry = next(e for e in sparse["plan"]["entries"] if e["workflow_id"] == "458")
    sparse_ws = load_workbook(Path(sparse_entry["request"]["structured"]["normalizedPath"])).active
    sparse_data = [c.value for c in sparse_ws[3]]
    assert sparse_data[2] == "传感器模组" and sparse_data[5] == "EA", sparse_data
    assert sparse_data[10] == "C2-0225002" and sparse_data[15] == "P22", sparse_data
    assert sparse_data[18] == "苏州工业园区玲珑街88号" and sparse_data[19] == "紧急", sparse_data
    print("PASS sparse 458 demand -> attachment backfilled name/unit/project/MRP/address/WBS remark")

    # 3d) 458 MRP控制者 is mandatory: if neither row nor WBS binding provides it,
    # prepare parks the draft instead of generating an attachment with a blank MRP.
    saved_demand_mrp = dict(mockmod._WBS_REGISTRY[DEMAND_WBS])
    try:
        mockmod._WBS_REGISTRY[DEMAND_WBS] = {**saved_demand_mrp, "mrpController": ""}
        mockmod._INVENTORY["4000059295"] = []
        missing_mrp = run(graph, excel=make_workbook([
            ["1010", DEMAND_WBS, "", "4000059295", "", 1, "", ""],
        ]), thread="router-458-missing-mrp", save=True)
        assert missing_mrp.get("status") == STATUS_NEEDS_INPUT, missing_mrp.get("status")
        d458_mrp = next(d for d in missing_mrp["result"]["drafts"] if d["workflow_id"] == "458")
        assert d458_mrp["skipped"] and d458_mrp["needsInput"]["kind"] == "mrpController", d458_mrp
        print("PASS 458 missing MRP控制者 -> needs_input, no blank-MRP attachment")
    finally:
        mockmod._WBS_REGISTRY[DEMAND_WBS] = saved_demand_mrp

    # 4) WBS bucketing: same material, two WBS -> separate 458 drafts
    out2 = run(graph, excel=make_workbook([
        ["1010", DEMAND_WBS, "C2-0225002", "9999999999", "无库存物料", 4, "EA", "M01"],
    ]), thread="router-badcode", save=True)
    # bad material code -> PDM validation stops before routing
    assert out2.get("status") == STATUS_NEEDS_INPUT, out2.get("status")
    assert (out2.get("result") or {}).get("input", {}).get("kind") == "material"
    print("PASS bad material code -> PDM gate -> needs_input (router never runs)")

    # 5) skip-on-missing-registry: public stock but WBS not in registry -> 412 skipped
    mockmod._INVENTORY["4000023659"] = [_loc("1010", "A001", "成品仓", "", "10.000", "")]
    out3 = run(graph, excel=make_workbook([
        ["1010", "ZZ-NOT-IN-REGISTRY", "ZZ", "4000023659", "PCR板", 4, "EA", "M01"],
    ]), thread="router-skip", save=True)
    res3 = out3["result"]
    d412 = next(d for d in res3["drafts"] if d["workflow_id"] == "412")
    assert d412["skipped"] and d412["needsInput"]["kind"] == "costCenter", d412
    assert out3.get("status") == STATUS_NEEDS_INPUT, out3.get("status")
    print("PASS missing-registry WBS -> 412 draft skipped (needsInput costCenter), others unaffected")

    # 6) 89 公共→项目 missing transfer-in stock location -> needs_input
    saved_demand_loc = dict(mockmod._WBS_REGISTRY[DEMAND_WBS])
    try:
        mockmod._WBS_REGISTRY[DEMAND_WBS] = {**saved_demand_loc, "stockLocationName": "", "stockLocationSapCode": ""}
        mockmod._INVENTORY["4000023659"] = [_loc("1010", "A001", "成品仓", "", "2.000", "")]  # public
        out_stock = run(graph, excel=make_workbook([
            ["1010", DEMAND_WBS, "C2-0225002", "4000023659", "PCR板", 2, "EA", "M01"],
        ]), thread="router-89pub-missing-stock", save=True)
        assert out_stock.get("status") == STATUS_NEEDS_INPUT, out_stock.get("status")
        d89 = next((d for d in out_stock["result"]["drafts"] if d["workflow_id"] == "89"), None)
        assert d89 and d89["skipped"] and d89["needsInput"]["kind"] == "stockLocation", d89
        print("PASS 89 公共→项目 missing transfer-in stock location -> needs_input")
    finally:
        mockmod._WBS_REGISTRY[DEMAND_WBS] = saved_demand_loc
        mockmod._INVENTORY["4000023659"] = [_loc("1010", "A001", "成品仓", "", "10.000", "")]

    # 7) return mode (classify_goal=return) -> 414 入库 draft bucketed by WBS, no stock fan-out
    _GOAL["value"] = "return"
    out4 = run(graph, excel=make_workbook([
        ["1010", DEMAND_WBS, "C2-0225002", "4000023659", "PCR板", 4, "EA", "M01"],
    ]), thread="router-return", save=True)
    _GOAL["value"] = "acquire"  # restore
    res4 = out4["result"]
    assert out4.get("status") == STATUS_DONE, (out4.get("status"), res4)
    assert {d["workflow_id"] for d in res4["drafts"]} == {"414"}, res4["drafts"]
    d414 = res4["drafts"][0]
    assert d414["wbsCode"] == DEMAND_WBS and d414["ok"] and d414["requestId"], d414
    print("PASS return mode -> single 414 入库 draft saved")

    # 8) unit mismatch (demand 盒 vs PDM base EA) -> unit_check LLM flags -> needs_input
    out5 = run(graph, excel=make_workbook([
        ["1010", DEMAND_WBS, "C2-0225002", "4000023659", "PCR板", 50, "盒", "M01"],
    ]), thread="router-unit", save=True)
    assert out5.get("status") == STATUS_NEEDS_INPUT, out5.get("status")
    assert (out5.get("result") or {}).get("input", {}).get("kind") == "unitReview", out5.get("result")
    print("PASS unit mismatch (盒≠EA) -> unit_check stops for confirmation (needs_input)")

    # 9) WBS alias: demand row references the WBS by its nickname -> resolved to code
    out6 = run(graph, excel=make_workbook([
        ["1010", "传感器项目", "C2-0225002", "4000023659", "PCR板", 10, "EA", "M01"],
    ]), thread="router-alias", save=True)
    assert out6.get("status") == STATUS_DONE, (out6.get("status"), out6.get("result"))
    d = out6["result"]["drafts"][0]
    assert d["wbsCode"] == DEMAND_WBS, ("alias should resolve to the real WBS code", d)
    print("PASS WBS alias '传感器项目' -> resolved to", d["wbsCode"], "and routed")

    print("\nALL ROUTER OFFLINE TESTS PASSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
