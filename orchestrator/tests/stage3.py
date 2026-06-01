"""Stage 3 offline tests: register the 412/414/458 workflows as plug-in flows
and drive each through the full graph against MockExecutor (no network, no
DeepSeek, no Edge). For every workflow we:

  1. build a tiny sample workbook matching that workflow's *_excel.py columns,
  2. parse it via the intake parser (intake_parsers.py),
  3. run the graph with run_workflow(workflow_id=..., executor=MockExecutor(),
     in-memory checkpointer) and assert status 'done' (--save => a mock
     requestId) OR an expected needs_input.

Mirrors tests/stage2.py structure.

Run: orchestrator/.venv/bin/python orchestrator/tests/stage3.py
"""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
os.environ["EXECUTOR"] = "mock"
os.environ["DEEPSEEK_API_KEY"] = ""  # force heuristic (offline gate; empty defeats .env dotenv reload)

from openpyxl import Workbook  # noqa: E402

from oa_orchestrator.config import get_settings  # noqa: E402
from oa_orchestrator.executors.mock import MockExecutor  # noqa: E402
from oa_orchestrator.graph import build_graph, make_checkpointer  # noqa: E402
from oa_orchestrator.intake_parsers import (parse_inbound, parse_outbound,  # noqa: E402
                                            parse_purchase)
from oa_orchestrator.runner import run_workflow  # noqa: E402
from oa_orchestrator.state import STATUS_DONE, STATUS_NEEDS_INPUT  # noqa: E402

CATALOG_CODE = "4000023659"  # exists + 启用 in MockExecutor catalog


# --------------------------------------------------------------------------- #
# Sample workbook builders (match each scripts/*_excel.py required columns)
# --------------------------------------------------------------------------- #
def make_outbound_workbook() -> str:
    wb = Workbook()
    main = wb.active
    main.title = "项目需求填写界面"
    main.append(["需求工厂代码", "WBS编码", "项目定义", "MRP控制者",
                 "物料编码", "物料名称", "需求数量", "基本计量单位"])  # row 1
    main.append(["说明"] * 8)                                          # row 2
    main.append(["1000", "C2-0225002.06.01", "C2-0225002", "M01",
                 CATALOG_CODE, "测试物料", 5, "EA"])                    # row 3
    mrp = wb.create_sheet("MRP控制者")
    mrp.append(["MRP编码", "MRP描述"])
    mrp.append(["M01", "ACRO"])
    cost = wb.create_sheet("成本中心")
    cost.append(["成本中心", "公司", "短文本"])
    cost.append(["CC1000", "1000", "产品开发部（ACRO）"])
    p = Path(tempfile.mkdtemp()) / "outbound.xlsx"
    wb.save(p)
    return str(p)


def make_inbound_workbook(code: str = CATALOG_CODE) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "项目需求填写界面"
    ws.append(["需求工厂代码", "WBS编码", "项目定义", "物料编码",
               "物料名称", "需求数量", "基本计量单位", "MRP控制者"])  # row 1
    ws.append(["说明"] * 8)                                            # row 2
    ws.append(["1000", "C2-0225002.06.01", "C2-0225002", code,
               "测试物料", 5, "EA", "M01"])                            # row 3
    p = Path(tempfile.mkdtemp()) / "inbound.xlsx"
    wb.save(p)
    return str(p)


def make_purchase_workbook() -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["需求日期", "项目定义", "WBS编码", "需求工厂代码"])      # row 1
    ws.append(["说明"] * 4)                                            # row 2
    ws.append(["20260101", "C2-0225002", "C2-0225002.06.01", "1000"])  # row 3
    p = Path(tempfile.mkdtemp()) / "purchase.xlsx"
    wb.save(p)
    return str(p)


def fresh_runtime():
    settings = get_settings()
    settings.ensure_runtime_dir()
    executor = MockExecutor()
    graph = build_graph(executor, checkpointer=make_checkpointer(None))  # in-memory, isolated
    return settings, executor, graph


# --------------------------------------------------------------------------- #
# Tests
# --------------------------------------------------------------------------- #
def test_outbound() -> None:
    xlsx = make_outbound_workbook()

    # 1) intake parse
    structured = parse_outbound(xlsx, "产品开发部ACRO")
    assert structured["ok"] is True
    assert structured["demandFactoryCode"] == "1000", structured["demandFactoryCode"]
    assert structured["wbsCode"] == "C2-0225002.06.01"
    assert structured["costCenter"]["costCenterCode"] == "CC1000", structured["costCenter"]
    assert len(structured["materialRows"]) == 1
    print("PASS 412 intake parse:", structured["costCenter"]["searchName"])

    # 2) graph -> done + mock requestId (department travels via profile)
    settings, executor, graph = fresh_runtime()
    res = run_workflow(request="物资出库", excel_path=xlsx, thread_id="s3-412",
                       save=True, workflow_id="412",
                       profile={"user_id": "bob", "department": "产品开发部ACRO"},
                       graph=graph, executor=executor, settings=settings)
    assert res["status"] == STATUS_DONE, res["status"]
    assert res["requestId"] == "MOCK-412-0001", res
    assert res["ok"] is True, res
    print("PASS 412 outbound -> status=done requestId=", res["requestId"])

    # 3) missing cost center -> executor needs_input (kind=costCenter). The
    # graph path above already proves wiring; validate the mock branch directly.
    _settings, executor, _graph = fresh_runtime()
    from oa_orchestrator.schemas import OutboundFillRequest
    bad = dict(structured)
    bad["costCenter"] = {}
    out = executor.fill_outbound(OutboundFillRequest(structured=bad, save=True))
    assert out.needsInput is True and out.input["kind"] == "costCenter", out
    print("PASS 412 missing cost center -> needsInput kind=costCenter")


def test_inbound() -> None:
    xlsx = make_inbound_workbook()

    # 1) intake parse
    structured = parse_inbound(xlsx)
    assert structured["ok"] is True
    assert structured["demandFactoryCode"] == "1000"
    assert structured["quantityByMaterialCode"].get(CATALOG_CODE) == "5", structured["quantityByMaterialCode"]
    print("PASS 414 intake parse:", structured["quantityByMaterialCode"])

    # 2) request carries a stock location ("设备零件仓 D002") -> done + requestId
    settings, executor, graph = fresh_runtime()
    res = run_workflow(request="项目退料入库到设备零件仓 D002", excel_path=xlsx,
                       thread_id="s3-414", save=True, workflow_id="414",
                       graph=graph, executor=executor, settings=settings)
    assert res["status"] == STATUS_DONE, res["status"]
    assert res["requestId"] == "MOCK-414-0001", res
    print("PASS 414 inbound (with stock location) -> status=done requestId=", res["requestId"])

    # 3) no stock location -> executor needs_input (kind=stockLocation)
    settings, executor, graph = fresh_runtime()
    res = run_workflow(request="项目退料入库", excel_path=xlsx,
                       thread_id="s3-414-nostock", save=True, workflow_id="414",
                       graph=graph, executor=executor, settings=settings)
    assert res["status"] == STATUS_NEEDS_INPUT, res["status"]
    assert (res["result"] or {}).get("input", {}).get("kind") == "stockLocation", res["result"]
    print("PASS 414 missing stock location -> status=needs_input kind=stockLocation")


def test_purchase() -> None:
    xlsx = make_purchase_workbook()

    # 1) intake parse + attachment normalization
    structured = parse_purchase(xlsx)
    assert structured["ok"] is True
    assert structured["demandFactoryCode"] == "1000"
    assert structured["demandCompanyName"] == "北京镁伽机器人科技有限公司"
    assert Path(structured["normalizedPath"]).exists(), structured["normalizedPath"]
    print("PASS 458 intake parse + normalized attachment:", Path(structured["normalizedPath"]).name)

    # 2) graph -> done + mock requestId (458 is attachment-driven, no materials)
    settings, executor, graph = fresh_runtime()
    res = run_workflow(request="采购申请", excel_path=xlsx, thread_id="s3-458",
                       save=True, workflow_id="458",
                       graph=graph, executor=executor, settings=settings)
    assert res["status"] == STATUS_DONE, res["status"]
    assert res["requestId"] == "MOCK-458-0001", res
    print("PASS 458 purchase -> status=done requestId=", res["requestId"])

    # 3) missing attachment -> executor needs_input (kind=attachment)
    settings, executor, graph = fresh_runtime()
    from oa_orchestrator.schemas import PurchaseFillRequest
    bad = dict(structured)
    bad.pop("normalizedPath", None)
    out = executor.fill_purchase(PurchaseFillRequest(structured=bad, save=True))
    assert out.needsInput is True and out.input["kind"] == "attachment", out
    print("PASS 458 missing attachment -> needsInput kind=attachment")


def main() -> int:
    test_outbound()
    test_inbound()
    test_purchase()
    print("\nALL STAGE 3 OFFLINE TESTS PASSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
